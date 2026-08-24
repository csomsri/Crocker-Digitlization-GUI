from __future__ import annotations

import json
import math
from collections.abc import Callable
from pathlib import Path
from typing import Any

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QCheckBox,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from python.app.PageShell import DetailPage
from python.app.widgets.MagneticFieldWidgets import CHANNEL_NAMES


Transform = dict[str, Any]
ChannelEntry = dict[str, Any]
TransformSummary = dict[str, QLabel]


class ScalingPage(DetailPage):
    CONFIG_RELATIVE_PATH = Path("config") / "trim_coil_scaling.json"
    CHANNEL_KEYS = [f"ch{index}" for index in range(1, 13)] + ["main_magnet", "centering_beam"]
    POINT_COLUMNS = ("Input", "Output")
    TABLE_ROW_HEIGHT = 40
    SUMMARY_ROW_HEIGHT = 68
    EDITOR_CONTROL_WIDTH = 450
    EDITOR_TABLE_GAP = 48
    EDITOR_TABLE_WIDTH = 240
    EDITOR_BODY_HEIGHT = 248
    EDITOR_SECTION_HEIGHT = 306
    EDITOR_SUMMARY_WIDTHS = (76, 82, 124)
    COLUMNS = (
        "Channel",
        "Enabled",
        "Raw to Eng",
        "Eng to Raw",
    )

    def __init__(
        self,
        go_back: Callable[[], None],
        apply_live_scaling: Callable[[dict[str, object]], bool] | None = None,
    ) -> None:
        super().__init__(
            "Scaling",
            "Scaling and calibration setup",
            "Back to Configuration",
            go_back,
        )

        self.config_path = Path(__file__).resolve().parents[3] / self.CONFIG_RELATIVE_PATH
        self._apply_live_scaling = apply_live_scaling
        self.enabled_boxes: list[QCheckBox] = []
        self.raw_to_eng_labels: list[TransformSummary] = []
        self.eng_to_raw_labels: list[TransformSummary] = []
        self.channel_entries: list[ChannelEntry] = []
        self.direction_widgets: dict[str, dict[str, object]] = {}
        self.selected_channel_index = 0
        self._loading = False
        self._dirty = False
        self._editor_loading = False

        _, workspace = self.add_workspace()
        workspace.setContentsMargins(10, 8, 10, 10)
        workspace.setSpacing(10)

        heading = QLabel("TRIM COIL SCALING")
        heading.setObjectName("settingsHeading")
        workspace.addWidget(heading)

        summary = QFrame()
        summary.setObjectName("displayModePanel")
        summary_layout = QGridLayout(summary)
        summary_layout.setContentsMargins(14, 10, 14, 10)
        summary_layout.setSpacing(8)
        self.enabled_summary = QLabel("0 / 14\nENABLED")
        self.enabled_summary.setObjectName("scalingSummary")
        self.validation_summary = QLabel("READY\nVALIDATION")
        self.validation_summary.setObjectName("scalingSummary")
        self.change_summary = QLabel("SAVED\nSTATE")
        self.change_summary.setObjectName("scalingSummary")
        summary_layout.addWidget(self.enabled_summary, 0, 0)
        summary_layout.addWidget(self.validation_summary, 0, 1)
        summary_layout.addWidget(self.change_summary, 0, 2)
        workspace.addWidget(summary)

        self.status_label = QLabel()
        self.status_label.setObjectName("settingsDescription")
        self.status_label.setWordWrap(True)
        workspace.addWidget(self.status_label)

        self.table = QTableWidget(len(CHANNEL_NAMES), len(self.COLUMNS))
        self.table.setObjectName("scalingTable")
        self.table.setHorizontalHeaderLabels(self.COLUMNS)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setDefaultSectionSize(self.TABLE_ROW_HEIGHT)
        self.table.setMinimumWidth(860)
        self.table.itemSelectionChanged.connect(self._select_table_row)

        self._build_rows()

        editor = self._build_transform_editor()
        calibration_workspace = QHBoxLayout()
        calibration_workspace.setSpacing(12)
        calibration_workspace.addWidget(self.table, 7)
        calibration_workspace.addWidget(editor, 4, alignment=Qt.AlignmentFlag.AlignTop)
        workspace.addLayout(calibration_workspace, 1)

        actions = QHBoxLayout()
        for label, handler in (
            ("Reload Saved", self._load_scaling),
            ("Import Curves", self._import_scaling_file),
            ("Apply Draft", self._apply_draft_scaling),
            ("Save + Apply", self._save_scaling),
            ("Reset Identity", self._reset_identity),
        ):
            button = QPushButton(label)
            button.setObjectName("fieldBulk")
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            button.clicked.connect(handler)
            actions.addWidget(button)
        actions.addSpacing(18)
        for label, enabled in (("Enable All", True), ("Disable All", False)):
            button = QPushButton(label)
            button.setObjectName("fieldBulk")
            button.clicked.connect(
                lambda checked=False, value=enabled: self._set_all_enabled(value)
            )
            actions.addWidget(button)
        actions.addStretch(1)
        workspace.addLayout(actions)

        self._load_scaling()

    def _build_rows(self) -> None:
        for row, channel in enumerate(CHANNEL_NAMES):
            name_item = QTableWidgetItem(channel)
            name_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            name_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 0, name_item)

            enabled = QCheckBox()
            enabled.setObjectName("toggleRow")
            enabled.setToolTip(f"Enable scaling for {channel}")
            enabled.toggled.connect(self._mark_dirty)
            self.table.setCellWidget(row, 1, self._centered_widget(enabled))
            self.enabled_boxes.append(enabled)

            raw_widget, raw_labels = self._make_transform_summary_widget()
            self.table.setCellWidget(row, 2, raw_widget)
            self.raw_to_eng_labels.append(raw_labels)

            eng_widget, eng_labels = self._make_transform_summary_widget()
            self.table.setCellWidget(row, 3, eng_widget)
            self.eng_to_raw_labels.append(eng_labels)

    def _make_transform_summary_widget(
        self,
        height: int | None = None,
        widths: tuple[int, int, int] | None = None,
    ) -> tuple[QFrame, TransformSummary]:
        summary_height = height if height is not None else self.TABLE_ROW_HEIGHT - 6
        compact = height is None or summary_height <= self.TABLE_ROW_HEIGHT
        kind_width, primary_width, secondary_width = widths or (
            62 if compact else 88,
            114 if compact else 138,
            114 if compact else 138,
        )
        frame = QFrame()
        frame.setObjectName("scalingTransformSummary")
        frame.setFixedHeight(summary_height)
        frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(8 if compact else 12, 2 if compact else 6, 8 if compact else 12, 2 if compact else 6)
        layout.setSpacing(8 if compact else 12)

        summary_object_prefix = "scalingEditorTransform" if widths is not None else "scalingTransform"

        kind = QLabel()
        kind.setObjectName(f"{summary_object_prefix}Kind")
        kind.setAlignment(Qt.AlignmentFlag.AlignCenter)
        kind.setFixedWidth(kind_width)
        layout.addWidget(kind)

        primary = QLabel()
        primary.setObjectName(f"{summary_object_prefix}Value")
        primary.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        primary.setFixedWidth(primary_width)
        layout.addWidget(primary)

        secondary = QLabel()
        secondary.setObjectName(f"{summary_object_prefix}Value")
        secondary.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        secondary.setFixedWidth(secondary_width)
        layout.addWidget(secondary)
        layout.addStretch(1)

        return frame, {"kind": kind, "primary": primary, "secondary": secondary}

    def _build_transform_editor(self) -> QFrame:
        editor = QFrame()
        editor.setObjectName("displayModePanel")
        editor.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        editor.setMinimumWidth(760)
        editor.setMaximumWidth(980)
        editor.setMinimumHeight((self.TABLE_ROW_HEIGHT * len(CHANNEL_NAMES)) + 72)
        editor_layout = QGridLayout(editor)
        editor_layout.setContentsMargins(14, 10, 14, 10)
        editor_layout.setHorizontalSpacing(12)
        editor_layout.setVerticalSpacing(10)
        editor_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.editor_title = QLabel("SELECT A CHANNEL")
        self.editor_title.setObjectName("scalingEditorTitle")
        self.editor_title.setFixedHeight(38)
        editor_layout.addWidget(self.editor_title, 0, 0, 1, 2)

        editor_layout.addWidget(
            self._build_direction_section("raw_to_eng", "Raw to Eng"),
            1,
            0,
            1,
            2,
            alignment=Qt.AlignmentFlag.AlignTop,
        )
        editor_layout.addWidget(
            self._build_direction_section("eng_to_raw", "Eng to Raw"),
            2,
            0,
            1,
            2,
            alignment=Qt.AlignmentFlag.AlignTop,
        )
        return editor

    def _build_direction_section(self, direction: str, title: str) -> QFrame:
        section = QFrame()
        section.setObjectName("scalingConversionCard")
        section.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        section.setMinimumHeight(self.EDITOR_SECTION_HEIGHT)
        section.setMaximumHeight(self.EDITOR_SECTION_HEIGHT)
        section_layout = QGridLayout(section)
        section_layout.setContentsMargins(12, 10, 12, 10)
        section_layout.setHorizontalSpacing(10)
        section_layout.setVerticalSpacing(8)
        section_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        section_layout.setColumnMinimumWidth(0, self.EDITOR_CONTROL_WIDTH)
        section_layout.setColumnMinimumWidth(1, self.EDITOR_TABLE_GAP)
        section_layout.setColumnMinimumWidth(2, self.EDITOR_TABLE_WIDTH)
        section_layout.setColumnStretch(0, 0)
        section_layout.setColumnStretch(1, 0)
        section_layout.setColumnStretch(2, 0)
        section_layout.setColumnStretch(3, 1)

        heading = QLabel(title)
        heading.setObjectName("scalingSectionTitle")
        heading.setFixedHeight(30)
        section_layout.addWidget(heading, 0, 0, 1, 4)

        left_panel = QFrame()
        left_panel.setObjectName("scalingCalibrationStack")
        left_panel.setFixedSize(self.EDITOR_CONTROL_WIDTH, self.EDITOR_BODY_HEIGHT)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        formula_frame, formula_summary = self._make_transform_summary_widget(
            self.SUMMARY_ROW_HEIGHT,
            self.EDITOR_SUMMARY_WIDTHS,
        )
        formula_frame.setObjectName("scalingFormula")
        formula_frame.setFixedWidth(350)
        left_layout.addWidget(formula_frame, alignment=Qt.AlignmentFlag.AlignLeft)

        controls_panel = QFrame()
        controls_panel.setObjectName("scalingControlsPanel")
        controls_panel.setFixedWidth(self.EDITOR_CONTROL_WIDTH)
        controls_layout = QGridLayout(controls_panel)
        controls_layout.setContentsMargins(0, 0, 0, 0)
        controls_layout.setHorizontalSpacing(8)
        controls_layout.setVerticalSpacing(7)
        controls_layout.setColumnMinimumWidth(0, 64)
        controls_layout.setColumnStretch(0, 0)
        controls_layout.setColumnStretch(1, 1)

        type_label = QLabel("TYPE")
        type_label.setObjectName("scalingFieldLabel")
        controls_layout.addWidget(type_label, 0, 0)
        type_toggle = QFrame()
        type_toggle.setObjectName("scalingTypeToggle")
        type_toggle.setFixedWidth(170)
        type_toggle_layout = QHBoxLayout(type_toggle)
        type_toggle_layout.setContentsMargins(2, 2, 2, 2)
        type_toggle_layout.setSpacing(2)

        type_group = QButtonGroup(type_toggle)
        type_group.setExclusive(True)
        for label, transform_type in (("Linear", "linear"), ("Curve", "curve")):
            button = QPushButton(label)
            button.setObjectName("scalingTypeButton")
            button.setCheckable(True)
            button.setProperty("transformType", transform_type)
            button.clicked.connect(
                lambda checked=False, selected_direction=direction: self._commit_editor_transform(selected_direction)
            )
            type_group.addButton(button)
            type_toggle_layout.addWidget(button)
        controls_layout.addWidget(type_toggle, 0, 1, alignment=Qt.AlignmentFlag.AlignLeft)

        linear_panel = QFrame()
        linear_panel.setObjectName("scalingLinearPanel")
        linear_panel.setFixedWidth(356)
        linear_layout = QGridLayout(linear_panel)
        linear_layout.setContentsMargins(0, 0, 0, 0)
        linear_layout.setHorizontalSpacing(14)
        linear_layout.setVerticalSpacing(6)

        gain_label = QLabel("GAIN")
        gain_label.setObjectName("scalingFieldLabel")
        linear_layout.addWidget(gain_label, 0, 0)
        gain_spin = self._make_transform_spinbox()
        gain_spin.setObjectName("scalingInput")
        gain_spin.setFixedWidth(170)
        gain_spin.valueChanged.connect(
            lambda _value, selected_direction=direction: self._commit_editor_transform(selected_direction)
        )
        linear_layout.addWidget(gain_spin, 1, 0)

        offset_label = QLabel("OFFSET")
        offset_label.setObjectName("scalingFieldLabel")
        linear_layout.addWidget(offset_label, 0, 1)
        offset_spin = self._make_transform_spinbox()
        offset_spin.setObjectName("scalingInput")
        offset_spin.setFixedWidth(170)
        offset_spin.valueChanged.connect(
            lambda _value, selected_direction=direction: self._commit_editor_transform(selected_direction)
        )
        linear_layout.addWidget(offset_spin, 1, 1)
        controls_layout.addWidget(linear_panel, 1, 1, alignment=Qt.AlignmentFlag.AlignLeft)

        sample_label = QLabel("CHECK")
        sample_label.setObjectName("scalingFieldLabel")
        controls_layout.addWidget(sample_label, 2, 0)
        check_panel = QFrame()
        check_panel.setObjectName("scalingCheckPanel")
        check_panel.setFixedWidth(354)
        check_layout = QHBoxLayout(check_panel)
        check_layout.setContentsMargins(0, 0, 0, 0)
        check_layout.setSpacing(14)
        sample_spin = self._make_transform_spinbox()
        sample_spin.setObjectName("scalingInput")
        sample_spin.setFixedWidth(170)
        sample_spin.valueChanged.connect(
            lambda _value, selected_direction=direction: self._refresh_preview(selected_direction)
        )
        check_layout.addWidget(sample_spin)
        result_label = QLabel("-> 0")
        result_label.setObjectName("scalingResult")
        result_label.setFixedWidth(170)
        check_layout.addWidget(result_label)
        controls_layout.addWidget(check_panel, 2, 1, alignment=Qt.AlignmentFlag.AlignLeft)

        linear_preview_panel = QFrame()
        linear_preview_panel.setObjectName("scalingLinearPreviewPanel")
        linear_preview_panel.setFixedSize(self.EDITOR_TABLE_WIDTH, self.EDITOR_BODY_HEIGHT)
        linear_preview_layout = QVBoxLayout(linear_preview_panel)
        linear_preview_layout.setContentsMargins(12, 12, 12, 12)
        linear_preview_layout.setSpacing(8)
        linear_preview_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        input_title = QLabel("INPUT")
        input_title.setObjectName("scalingPreviewCaption")
        input_title.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        linear_preview_layout.addWidget(input_title)

        preview_input = QLabel("0")
        preview_input.setObjectName("scalingPreviewValue")
        preview_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        preview_input.setFixedSize(216, 58)
        linear_preview_layout.addWidget(preview_input)

        preview_arrow = QLabel("v")
        preview_arrow.setObjectName("scalingPreviewArrow")
        preview_arrow.setAlignment(Qt.AlignmentFlag.AlignCenter)
        preview_arrow.setFixedHeight(26)
        linear_preview_layout.addWidget(preview_arrow)

        output_title = QLabel("OUTPUT")
        output_title.setObjectName("scalingPreviewCaption")
        output_title.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        linear_preview_layout.addWidget(output_title)

        preview_output = QLabel("0")
        preview_output.setObjectName("scalingPreviewValue")
        preview_output.setAlignment(Qt.AlignmentFlag.AlignCenter)
        preview_output.setFixedSize(216, 58)
        linear_preview_layout.addWidget(preview_output)

        curve_panel = QFrame()
        curve_panel.setObjectName("scalingCurvePanel")
        curve_panel.setFixedSize(self.EDITOR_TABLE_WIDTH, self.EDITOR_BODY_HEIGHT)
        curve_layout = QVBoxLayout(curve_panel)
        curve_layout.setContentsMargins(0, 0, 0, 0)
        curve_layout.setSpacing(8)

        points_table = QTableWidget(0, len(self.POINT_COLUMNS))
        points_table.setObjectName("scalingPointsTable")
        points_table.setHorizontalHeaderLabels(self.POINT_COLUMNS)
        points_table.verticalHeader().setVisible(False)
        points_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        points_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        points_table.verticalHeader().setDefaultSectionSize(32)
        points_table.setFixedWidth(self.EDITOR_TABLE_WIDTH)
        points_table.setFixedHeight(self.EDITOR_BODY_HEIGHT)
        points_table.itemChanged.connect(
            lambda _item, selected_direction=direction: self._commit_editor_transform(selected_direction)
        )
        curve_layout.addWidget(points_table)

        point_actions_frame = QFrame()
        point_actions_frame.setObjectName("scalingPointActions")
        point_actions_frame.setFixedWidth(356)
        point_actions = QHBoxLayout(point_actions_frame)
        point_actions.setContentsMargins(0, 0, 0, 0)
        point_actions.setSpacing(8)
        for label, width, handler in (
            ("Add Point", 86, self._add_editor_point),
            ("Remove Selected", 122, self._remove_editor_points),
            ("Sort By Input", 108, self._sort_editor_points),
        ):
            button = QPushButton(label)
            button.setObjectName("scalingPointButton")
            button.setFixedWidth(width)
            button.clicked.connect(lambda checked=False, selected_direction=direction, selected_handler=handler: selected_handler(selected_direction))
            point_actions.addWidget(button)
        point_actions.addStretch(1)
        controls_layout.addWidget(point_actions_frame, 3, 1, alignment=Qt.AlignmentFlag.AlignLeft)
        left_layout.addWidget(controls_panel, alignment=Qt.AlignmentFlag.AlignLeft)
        left_layout.addStretch(1)

        section_layout.addWidget(left_panel, 1, 0, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        section_layout.addWidget(linear_preview_panel, 1, 2, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        section_layout.addWidget(curve_panel, 1, 2, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)

        self.direction_widgets[direction] = {
            "type_group": type_group,
            "formula_summary": formula_summary,
            "gain_label": gain_label,
            "gain_spin": gain_spin,
            "offset_label": offset_label,
            "offset_spin": offset_spin,
            "linear_panel": linear_panel,
            "sample_spin": sample_spin,
            "result_label": result_label,
            "check_panel": check_panel,
            "linear_preview_panel": linear_preview_panel,
            "preview_input": preview_input,
            "preview_output": preview_output,
            "curve_panel": curve_panel,
            "points_table": points_table,
            "point_actions_frame": point_actions_frame,
        }
        return section

    def _make_transform_spinbox(self) -> QDoubleSpinBox:
        spinbox = QDoubleSpinBox()
        spinbox.setDecimals(10)
        spinbox.setRange(-1.0e12, 1.0e12)
        spinbox.setSingleStep(0.01)
        spinbox.setKeyboardTracking(False)
        spinbox.setButtonSymbols(QDoubleSpinBox.ButtonSymbols.NoButtons)
        return spinbox

    def _centered_widget(self, child: QWidget) -> QWidget:
        frame = QFrame()
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addStretch(1)
        layout.addWidget(child)
        layout.addStretch(1)
        return frame

    def _identity_scaling(self) -> dict[str, object]:
        return {
            key: {
                "label": CHANNEL_NAMES[index],
                "enabled": False,
                "raw_to_eng": {"type": "linear", "gain": 1.0, "offset": 0.0},
                "eng_to_raw": {"type": "linear", "gain": 1.0, "offset": 0.0},
            }
            for index, key in enumerate(self.CHANNEL_KEYS)
        }

    def _load_scaling(self) -> None:
        scaling = self._identity_scaling()
        if self.config_path.exists():
            try:
                with self.config_path.open("r", encoding="utf-8") as handle:
                    data = json.load(handle)
                scaling = self._normalize_scaling(data)
                self.status_label.setText(f"Loaded {self.config_path}")
            except Exception as exc:
                self.status_label.setText(f"Could not load scaling file: {exc}")
        else:
            self.status_label.setText(f"No active scaling file. Save creates {self.config_path}")
        self._apply_scaling_to_table(scaling)
        self._set_clean()

    def _save_scaling(self) -> None:
        scaling = self._read_scaling_from_table()
        errors = self._validation_errors(scaling)
        if errors:
            self.status_label.setText("Cannot save: " + "; ".join(errors[:3]))
            self._refresh_summary()
            return
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        with self.config_path.open("w", encoding="utf-8") as handle:
            json.dump(scaling, handle, indent=2)
            handle.write("\n")
        enabled_count = self._enabled_scaling_count(scaling)
        live_status = ""
        if self._apply_live_scaling is not None:
            live_status = " Live backend updated." if self._apply_live_scaling(scaling) else " Live backend not active."
        self.status_label.setText(
            f"Saved {enabled_count}/{len(CHANNEL_NAMES)} enabled channels to {self.config_path}.{live_status}"
        )

        self._set_clean()

    def _apply_draft_scaling(self) -> None:
        scaling = self._read_scaling_from_table()
        errors = self._validation_errors(scaling)
        if errors:
            self.status_label.setText("Cannot apply draft: " + "; ".join(errors[:3]))
            self._refresh_summary()
            return
        if self._apply_live_scaling is None:
            self.status_label.setText("No live Field Ctrl backend is available.")
            return
        applied = self._apply_live_scaling(scaling)
        self.status_label.setText(
            "Draft applied to the live backend; it has not been saved."
            if applied else "Live Field Ctrl backend is not active."
        )

    def _import_scaling_file(self) -> None:
        path_text, _selected_filter = QFileDialog.getOpenFileName(
            self,
            "Import scaling curves",
            str(self.config_path.parent),
            "Scaling files (*.json *.xlsx);;JSON files (*.json);;Excel workbooks (*.xlsx)",
        )
        if not path_text:
            return

        path = Path(path_text)
        try:
            if path.suffix.lower() == ".json":
                with path.open("r", encoding="utf-8") as handle:
                    scaling = self._normalize_scaling(json.load(handle))
            elif path.suffix.lower() == ".xlsx":
                scaling = self._normalize_scaling(self._scaling_from_workbook(path))
            else:
                raise ValueError("import file must be .json or .xlsx")
        except Exception as exc:
            self.status_label.setText(f"Could not import scaling file: {exc}")
            return

        self._apply_scaling_to_table(scaling)
        self.status_label.setText(f"Imported scaling draft from {path}")
        self._mark_dirty()

    def _apply_file_scaling(self) -> None:
        if self._apply_live_scaling is None:
            self.status_label.setText("No live Field Ctrl backend is available.")
            return
        if not self.config_path.exists():
            self.status_label.setText(f"No active scaling file at {self.config_path}")
            return

        try:
            with self.config_path.open("r", encoding="utf-8") as handle:
                data = json.load(handle)
            scaling = self._normalize_scaling(data)
        except Exception as exc:
            self.status_label.setText(f"Could not apply scaling file: {exc}")
            return

        live_status = "Live backend updated." if self._apply_live_scaling(scaling) else "Live backend not active."
        self.status_label.setText(f"Applied {self.config_path}. {live_status}")

    def _reset_identity(self) -> None:
        self._apply_scaling_to_table(self._identity_scaling())
        self.status_label.setText("Reset table to identity scaling. Save to make it active.")
        self._mark_dirty()

    def _set_all_enabled(self, enabled: bool) -> None:
        for box in self.enabled_boxes:
            box.setChecked(enabled)
        self.status_label.setText(
            "All channels enabled in the draft."
            if enabled else "All channels disabled in the draft."
        )

    def _edit_transform(self, index: int, direction: str) -> None:
        self.selected_channel_index = index
        self.table.selectRow(index)
        self._load_editor_from_selection()

    def _select_table_row(self) -> None:
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            return
        self.selected_channel_index = selected[0].row()
        self._load_editor_from_selection()

    def _load_editor_from_selection(self) -> None:
        if not self.channel_entries:
            return

        self.selected_channel_index = max(0, min(self.selected_channel_index, len(self.channel_entries) - 1))
        entry = self.channel_entries[self.selected_channel_index]

        self._editor_loading = True
        self.editor_title.setText(f"{CHANNEL_NAMES[self.selected_channel_index]} CALIBRATION")
        for direction in ("raw_to_eng", "eng_to_raw"):
            transform = self._normalize_transform(entry.get(direction))
            widgets = self.direction_widgets[direction]
            type_group = widgets["type_group"]
            gain_spin = widgets["gain_spin"]
            offset_spin = widgets["offset_spin"]
            if isinstance(type_group, QButtonGroup):
                self._set_transform_type(type_group, str(transform.get("type", "linear")))
            if isinstance(gain_spin, QDoubleSpinBox):
                gain_spin.setValue(float(transform.get("gain", 1.0)))
            if isinstance(offset_spin, QDoubleSpinBox):
                offset_spin.setValue(float(transform.get("offset", 0.0)))
            self._set_editor_points(direction, self._points_from_transform(transform))
        self._editor_loading = False
        for direction in ("raw_to_eng", "eng_to_raw"):
            self._refresh_editor_state(direction)
            self._refresh_formula(direction)
            self._refresh_preview(direction)

    def _commit_editor_transform(self, direction: str, *_ignored) -> None:
        if self._editor_loading or not self.channel_entries:
            return
        widgets = self.direction_widgets[direction]
        type_group = widgets["type_group"]
        gain_spin = widgets["gain_spin"]
        offset_spin = widgets["offset_spin"]
        if not isinstance(type_group, QButtonGroup) or not isinstance(gain_spin, QDoubleSpinBox) or not isinstance(offset_spin, QDoubleSpinBox):
            return
        transform = self._normalize_transform(
            {
                "type": self._transform_type(type_group),
                "gain": gain_spin.value(),
                "offset": offset_spin.value(),
                "points": self._read_editor_points(direction),
            }
        )
        self.channel_entries[self.selected_channel_index][direction] = transform
        self._refresh_transform_labels()
        self._refresh_editor_state(direction)
        self._refresh_formula(direction)
        self._refresh_preview(direction)
        self._mark_dirty()

    def _refresh_formula(self, direction: str) -> None:
        if not self.channel_entries:
            return
        widgets = self.direction_widgets[direction]
        formula_summary = widgets["formula_summary"]
        if not isinstance(formula_summary, dict):
            return
        transform = self.channel_entries[self.selected_channel_index].get(direction)
        if not isinstance(transform, dict):
            return
        self._set_transform_summary(formula_summary, transform)

    def _refresh_preview(self, direction: str) -> None:
        if self._editor_loading or not self.channel_entries:
            return
        widgets = self.direction_widgets[direction]
        sample_spin = widgets["sample_spin"]
        result_label = widgets["result_label"]
        if not isinstance(sample_spin, QDoubleSpinBox) or not isinstance(result_label, QLabel):
            return
        transform = self.channel_entries[self.selected_channel_index].get(direction)
        if not isinstance(transform, dict):
            return
        try:
            result = self._apply_transform(sample_spin.value(), transform)
        except Exception:
            result_label.setText("-> invalid")
            return
        result_label.setText(f"-> {result:.10g}")
        self._refresh_linear_preview(direction)

    def _refresh_linear_preview(self, direction: str) -> None:
        widgets = self.direction_widgets[direction]
        preview_input = widgets.get("preview_input")
        preview_output = widgets.get("preview_output")
        sample_spin = widgets.get("sample_spin")
        if not isinstance(preview_input, QLabel) or not isinstance(preview_output, QLabel) or not isinstance(sample_spin, QDoubleSpinBox) or not self.channel_entries:
            return
        transform = self.channel_entries[self.selected_channel_index].get(direction)
        if not isinstance(transform, dict) or transform.get("type") == "curve":
            return

        input_value = sample_spin.value()
        output_value = self._apply_transform(input_value, transform)
        preview_input.setText(f"{input_value:.10g}")
        preview_output.setText(f"{output_value:.10g}")

    def _apply_transform(self, value: float, transform: Transform) -> float:
        if transform.get("type") == "curve":
            points = sorted(self._points_from_transform(transform), key=lambda point: point[0])
            if not points:
                return value
            if value <= points[0][0]:
                return points[0][1]
            if value >= points[-1][0]:
                return points[-1][1]
            for low, high in zip(points, points[1:]):
                if low[0] <= value <= high[0]:
                    span = high[0] - low[0]
                    if span == 0.0:
                        return low[1]
                    fraction = (value - low[0]) / span
                    return low[1] + fraction * (high[1] - low[1])
            return points[-1][1]
        return value * float(transform.get("gain", 1.0)) + float(transform.get("offset", 0.0))

    def _refresh_editor_state(self, direction: str) -> None:
        widgets = self.direction_widgets[direction]
        type_group = widgets["type_group"]
        curve_mode = isinstance(type_group, QButtonGroup) and self._transform_type(type_group) == "curve"
        for key in ("linear_panel",):
            widget = widgets[key]
            if isinstance(widget, QWidget):
                widget.setVisible(not curve_mode)
        for key in ("linear_preview_panel",):
            widget = widgets[key]
            if isinstance(widget, QWidget):
                widget.setVisible(not curve_mode)
        for key in ("curve_panel", "point_actions_frame"):
            widget = widgets[key]
            if isinstance(widget, QWidget):
                widget.setEnabled(curve_mode)
                widget.setVisible(curve_mode)

    def _transform_type(self, button_group: QButtonGroup) -> str:
        button = button_group.checkedButton()
        if button is None:
            return "linear"
        transform_type = button.property("transformType")
        return transform_type if isinstance(transform_type, str) else "linear"

    def _set_transform_type(self, button_group: QButtonGroup, transform_type: str) -> None:
        for button in button_group.buttons():
            button.setChecked(button.property("transformType") == transform_type)

    def _set_editor_points(self, direction: str, points: list[list[float]]) -> None:
        table = self.direction_widgets[direction]["points_table"]
        if not isinstance(table, QTableWidget):
            return
        table.setRowCount(0)
        if not points:
            points = [[0.0, 0.0], [1.0, 1.0]]
        for input_value, output_value in points:
            self._insert_editor_point(direction, float(input_value), float(output_value))

    def _insert_editor_point(self, direction: str, input_value: float, output_value: float) -> None:
        table = self.direction_widgets[direction]["points_table"]
        if not isinstance(table, QTableWidget):
            return
        row = table.rowCount()
        table.insertRow(row)
        for column, value in enumerate((input_value, output_value)):
            item = QTableWidgetItem(f"{value:.10g}")
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            table.setItem(row, column, item)

    def _add_editor_point(self, direction: str) -> None:
        self._insert_editor_point(direction, 0.0, 0.0)
        self._commit_editor_transform(direction)

    def _remove_editor_points(self, direction: str) -> None:
        table = self.direction_widgets[direction]["points_table"]
        if not isinstance(table, QTableWidget):
            return
        rows = sorted({index.row() for index in table.selectedIndexes()}, reverse=True)
        for row in rows:
            table.removeRow(row)
        self._commit_editor_transform(direction)

    def _sort_editor_points(self, direction: str) -> None:
        points = sorted(self._read_editor_points(direction), key=lambda point: point[0])
        self._editor_loading = True
        self._set_editor_points(direction, points)
        self._editor_loading = False
        self._commit_editor_transform(direction)

    def _read_editor_points(self, direction: str) -> list[list[float]]:
        table = self.direction_widgets[direction]["points_table"]
        if not isinstance(table, QTableWidget):
            return []
        points: list[list[float]] = []
        for row in range(table.rowCount()):
            input_item = table.item(row, 0)
            output_item = table.item(row, 1)
            if input_item is None or output_item is None:
                continue
            input_text = input_item.text().strip()
            output_text = output_item.text().strip()
            if not input_text or not output_text:
                continue
            points.append([float(input_text), float(output_text)])
        return points

    def _mark_dirty(self, *_ignored) -> None:
        if self._loading:
            return
        self._dirty = True
        self._refresh_summary()

    def _set_clean(self) -> None:
        self._dirty = False
        self._refresh_summary()

    def _validation_errors(self, scaling: dict[str, object]) -> list[str]:
        errors: list[str] = []
        for index, key in enumerate(self.CHANNEL_KEYS):
            entry = scaling.get(key)
            if not isinstance(entry, dict) or not bool(entry.get("enabled", True)):
                continue
            for direction, label in (("raw_to_eng", "raw to eng"), ("eng_to_raw", "eng to raw")):
                transform = entry.get(direction)
                if not isinstance(transform, dict):
                    errors.append(f"{CHANNEL_NAMES[index]} is missing {label}")
                    continue
                errors.extend(self._transform_errors(CHANNEL_NAMES[index], label, transform))
        return errors

    def _transform_errors(self, channel: str, label: str, transform: Transform) -> list[str]:
        transform_type = str(transform.get("type", "linear"))
        if transform_type == "curve":
            points = self._points_from_transform(transform)
            if len(points) < 2:
                return [f"{channel} {label} curve needs at least 2 points"]
            inputs = [point[0] for point in points]
            if any(not math.isfinite(value) for point in points for value in point):
                return [f"{channel} {label} curve has non-finite values"]
            if len(set(inputs)) != len(inputs):
                return [f"{channel} {label} curve has duplicate inputs"]
            return []

        gain = float(transform.get("gain", 1.0))
        offset = float(transform.get("offset", 0.0))
        if not math.isfinite(gain) or not math.isfinite(offset):
            return [f"{channel} {label} linear values must be finite"]
        if gain == 0.0:
            return [f"{channel} {label} has a zero gain"]
        return []

    def _refresh_summary(self) -> None:
        scaling = self._read_scaling_from_table()
        enabled_count = self._enabled_scaling_count(scaling)
        errors = self._validation_errors(scaling)
        self.enabled_summary.setText(f"{enabled_count} / {len(CHANNEL_NAMES)}\nENABLED")
        self.validation_summary.setText(
            f"{len(errors)} ISSUE{'S' if len(errors) != 1 else ''}\nVALIDATION"
            if errors else "PASSED\nVALIDATION"
        )
        self.validation_summary.setProperty("warning", bool(errors))
        self.validation_summary.style().unpolish(self.validation_summary)
        self.validation_summary.style().polish(self.validation_summary)
        self.change_summary.setText("UNSAVED\nSTATE" if self._dirty else "SAVED\nSTATE")

    def _normalize_scaling(self, data: object) -> dict[str, object]:
        scaling = self._identity_scaling()
        if not isinstance(data, dict):
            raise ValueError("scaling file must contain a JSON object")

        array_keys = {"raw_to_eng_gain", "raw_to_eng_offset", "eng_to_raw_gain", "eng_to_raw_offset", "enabled"}
        if array_keys.intersection(data):
            self._apply_array_scaling(scaling, data, array_keys)
            return scaling

        for index, key in enumerate(self.CHANNEL_KEYS):
            entry = data.get(key)
            if isinstance(entry, dict):
                scaling[key] = self._normalize_channel_entry(index, entry)
        return scaling

    def _apply_array_scaling(
        self,
        scaling: dict[str, object],
        data: dict[str, object],
        array_keys: set[str],
    ) -> None:
        arrays: dict[str, list[object]] = {}
        for key in array_keys:
            values = data.get(key)
            if values is None:
                continue
            if not isinstance(values, list) or len(values) != len(CHANNEL_NAMES):
                raise ValueError(f"{key} must contain {len(CHANNEL_NAMES)} entries")
            arrays[key] = values

        for index, channel_key in enumerate(self.CHANNEL_KEYS):
            entry = scaling[channel_key]
            if not isinstance(entry, dict):
                continue
            entry["enabled"] = bool(arrays.get("enabled", [False] * len(CHANNEL_NAMES))[index])
            entry["raw_to_eng"] = {
                "type": "linear",
                "gain": float(arrays.get("raw_to_eng_gain", [1.0] * len(CHANNEL_NAMES))[index]),
                "offset": float(arrays.get("raw_to_eng_offset", [0.0] * len(CHANNEL_NAMES))[index]),
            }
            entry["eng_to_raw"] = {
                "type": "linear",
                "gain": float(arrays.get("eng_to_raw_gain", [1.0] * len(CHANNEL_NAMES))[index]),
                "offset": float(arrays.get("eng_to_raw_offset", [0.0] * len(CHANNEL_NAMES))[index]),
            }

    def _normalize_channel_entry(self, index: int, entry: dict[str, object]) -> ChannelEntry:
        normalized: ChannelEntry = {
            "label": str(entry.get("label", CHANNEL_NAMES[index])),
            "enabled": bool(entry.get("enabled", True)),
            "raw_to_eng": self._normalize_transform(entry.get("raw_to_eng")),
            "eng_to_raw": self._normalize_transform(entry.get("eng_to_raw")),
        }
        return normalized

    def _normalize_transform(self, transform: object) -> Transform:
        if not isinstance(transform, dict):
            return {"type": "linear", "gain": 1.0, "offset": 0.0}

        transform_type = str(
            transform.get("type", "curve" if ("points" in transform or "curve" in transform) else "linear")
        ).strip().lower()
        if transform_type not in {"linear", "curve"}:
            transform_type = "linear"
        normalized: Transform = {
            "type": transform_type,
            "gain": float(transform.get("gain", 1.0)),
            "offset": float(transform.get("offset", 0.0)),
        }
        if transform_type == "curve":
            normalized["points"] = self._points_from_transform(transform)
        return normalized

    def _points_from_transform(self, transform: Transform) -> list[list[float]]:
        points = transform.get("points", transform.get("curve", []))
        parsed: list[list[float]] = []
        if not isinstance(points, list):
            return parsed
        for point in points:
            if isinstance(point, dict):
                if "input" in point and "output" in point:
                    parsed.append([float(point["input"]), float(point["output"])])
                elif "x" in point and "y" in point:
                    parsed.append([float(point["x"]), float(point["y"])])
                elif "raw" in point and "eng" in point:
                    parsed.append([float(point["raw"]), float(point["eng"])])
            elif isinstance(point, (list, tuple)) and len(point) >= 2:
                parsed.append([float(point[0]), float(point[1])])
        return parsed

    def _apply_scaling_to_table(self, scaling: dict[str, object]) -> None:
        self._loading = True
        self.channel_entries = []
        identity = self._identity_scaling()
        for index, key in enumerate(self.CHANNEL_KEYS):
            entry = scaling.get(key)
            if not isinstance(entry, dict):
                entry = identity[key]
            self.channel_entries.append(self._normalize_channel_entry(index, entry))
            self.enabled_boxes[index].setChecked(bool(self.channel_entries[index]["enabled"]))
        self._loading = False
        self._refresh_transform_labels()
        if self.channel_entries:
            self.selected_channel_index = min(self.selected_channel_index, len(self.channel_entries) - 1)
            self.table.selectRow(self.selected_channel_index)
            self._load_editor_from_selection()
        self._refresh_summary()

    def _read_scaling_from_table(self) -> dict[str, object]:
        scaling: dict[str, object] = {}
        for index, key in enumerate(self.CHANNEL_KEYS):
            entry = dict(self.channel_entries[index])
            entry["enabled"] = self.enabled_boxes[index].isChecked()
            scaling[key] = {
                "label": entry.get("label", CHANNEL_NAMES[index]),
                "enabled": entry["enabled"],
                "raw_to_eng": entry["raw_to_eng"],
                "eng_to_raw": entry["eng_to_raw"],
            }
        return scaling

    def _refresh_transform_labels(self) -> None:
        for index, entry in enumerate(self.channel_entries):
            self._set_transform_summary(self.raw_to_eng_labels[index], entry["raw_to_eng"])
            self._set_transform_summary(self.eng_to_raw_labels[index], entry["eng_to_raw"])

    def _set_transform_summary(self, labels: TransformSummary, transform: Transform) -> None:
        if transform.get("type") == "curve":
            points = self._points_from_transform(transform)
            labels["kind"].setText("CURVE")
            labels["kind"].setProperty("mode", "curve")
            labels["primary"].setText(f"{len(points)} POINTS")
            labels["secondary"].setText("INTERPOLATION")
            self._refresh_transform_summary_style(labels)
            return

        gain = float(transform.get("gain", 1.0))
        offset = float(transform.get("offset", 0.0))
        labels["kind"].setText("LINEAR")
        labels["kind"].setProperty("mode", "linear")
        labels["primary"].setText(f"SCALE {self._format_compact_number(gain)}")
        labels["secondary"].setText(f"OFFSET {self._format_compact_number(offset)}")
        self._refresh_transform_summary_style(labels)

    def _refresh_transform_summary_style(self, labels: TransformSummary) -> None:
        for label in labels.values():
            label.style().unpolish(label)
            label.style().polish(label)

    def _format_compact_number(self, value: float) -> str:
        if value == 0.0:
            return "0"
        absolute = abs(value)
        for suffix, scale in (("G", 1.0e9), ("M", 1.0e6), ("k", 1.0e3)):
            if absolute >= scale:
                return f"{value / scale:.5g}{suffix}"
        if absolute < 1.0e-3:
            return f"{value:.3e}"
        return f"{value:.6g}"

    def _enabled_scaling_count(self, scaling: dict[str, object]) -> int:
        count = 0
        for key in self.CHANNEL_KEYS:
            entry = scaling.get(key)
            if isinstance(entry, dict) and bool(entry.get("enabled", True)):
                count += 1
        return count

    def _scaling_from_workbook(self, path: Path) -> dict[str, object]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to import scaling_curves.xlsx") from exc

        workbook = load_workbook(path, data_only=True, read_only=True)
        scaling = self._identity_scaling()
        imported = 0
        for index, key in enumerate(self.CHANNEL_KEYS):
            if key not in workbook.sheetnames:
                continue
            sheet = workbook[key]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
            raw_index = self._find_header(headers, ("raw", "input", "x"))
            eng_index = self._find_header(headers, ("eng", "engineering", "output", "y"))
            if raw_index is None or eng_index is None:
                continue

            points: list[list[float]] = []
            for row in rows[1:]:
                if raw_index >= len(row) or eng_index >= len(row):
                    continue
                raw = row[raw_index]
                eng = row[eng_index]
                if raw is None or eng is None:
                    continue
                points.append([float(raw), float(eng)])
            if len(points) < 2:
                continue

            points = sorted(points, key=lambda point: point[0])
            inverse_points = sorted([[eng, raw] for raw, eng in points], key=lambda point: point[0])
            scaling[key] = {
                "label": CHANNEL_NAMES[index],
                "enabled": True,
                "raw_to_eng": {"type": "curve", "points": points},
                "eng_to_raw": {"type": "curve", "points": inverse_points},
            }
            imported += 1

        if imported == 0:
            raise ValueError("workbook did not contain matching channel sheets with raw/eng points")
        return scaling

    def _find_header(self, headers: list[str], names: tuple[str, ...]) -> int | None:
        for name in names:
            for index, header in enumerate(headers):
                if header == name or name in header:
                    return index
        return None
